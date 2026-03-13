"""Lanza N instancias paralelas de update_urls.py y fusiona los parciales al final."""
import subprocess
import os
import sys
import argparse
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
PYTHON    = PROJECT_ROOT / "python_portable" / "python.exe"
UPDATE_PY = PROJECT_ROOT / "scraper" / "update_urls.py"


def merge_partial_excels(excel_file: str, num_workers: int):
    """Fusiona los N parciales _updated_partial_wN.xlsx en el Excel final."""
    from openpyxl import load_workbook

    # Recopilar todos los registros actualizados de cada worker (keyed por URL)
    all_updated = {}
    for w in range(1, num_workers + 1):
        partial = excel_file.replace('.xlsx', f'_updated_partial_w{w}.xlsx')
        if not Path(partial).exists():
            print(f"  [!] Parcial worker {w} no encontrado: {partial}")
            continue
        wb = load_workbook(partial)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [c.value for c in ws[1]]
            if not headers or 'URL' not in headers:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(headers, row))
                if d.get('URL'):
                    all_updated[d['URL']] = d
        print(f"  [+] Parcial worker {w}: {len(all_updated)} registros acumulados")

    if not all_updated:
        print("[!] No se encontraron registros en los parciales. Abortando fusión.")
        return

    # Aplicar actualizaciones al Excel original y guardar como _updated.xlsx
    out = excel_file.replace('.xlsx', '_updated.xlsx')
    orig_wb = load_workbook(excel_file)
    updated_count = 0
    for sheet in orig_wb.sheetnames:
        ws = orig_wb[sheet]
        headers = [c.value for c in ws[1]]
        if not headers or 'URL' not in headers:
            continue
        url_col = headers.index('URL')
        for row in ws.iter_rows(min_row=2):
            url = row[url_col].value
            if url and url in all_updated:
                for i, header in enumerate(headers):
                    if header in all_updated[url]:
                        row[i].value = all_updated[url][header]
                updated_count += 1
    orig_wb.save(out)
    print(f"[✓] Fusionado: {updated_count} filas actualizadas → {out}")


def main():
    p = argparse.ArgumentParser(description="Lanzador paralelo para update_urls.py")
    p.add_argument("excel_file", help="Ruta al archivo Excel con las URLs")
    p.add_argument("--workers", type=int, default=5, help="Número de workers paralelos (default: 5)")
    p.add_argument("--mode", default="fast", choices=["fast", "stealth", "extra-stealth"],
                   help="Modo de scraping (default: fast)")
    p.add_argument("--resume", action="store_true", help="Reanudar desde journal existente")
    args = p.parse_args()

    if not Path(args.excel_file).exists():
        print(f"[ERROR] Archivo no encontrado: {args.excel_file}")
        sys.exit(1)

    print(f"[+] Iniciando update_urls paralelo con {args.workers} workers...")
    processes = []
    for worker_id in range(1, args.workers + 1):
        env = os.environ.copy()
        env["SCRAPER_WORKER_ID"]   = str(worker_id)
        env["SCRAPER_NUM_WORKERS"] = str(args.workers)
        cmd = [str(PYTHON), str(UPDATE_PY), args.excel_file, "--mode", args.mode]
        if args.resume:
            cmd.append("--resume")
        proc = subprocess.Popen(cmd, env=env)
        print(f"[+] Update Worker {worker_id} iniciado (PID {proc.pid})")
        processes.append((worker_id, proc))

    for worker_id, proc in processes:
        proc.wait()
        print(f"[✓] Update Worker {worker_id} terminado (código {proc.returncode})")

    print("[+] Fusionando resultados parciales...")
    merge_partial_excels(args.excel_file, args.workers)
    print("[✓] Proceso completo.")


if __name__ == "__main__":
    main()
