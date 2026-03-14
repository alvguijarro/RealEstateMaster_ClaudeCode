"""Flask-SocketIO web server for the Idealista scraper.

Provides REST API endpoints and WebSocket communication for real-time updates.
"""
from __future__ import annotations

import os
import sys
import asyncio
import threading
import json
import subprocess
import time
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, jsonify, request, send_file
from flask_socketio import SocketIO

# Add parent directory to path for idealista_scraper imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.scraper_wrapper import ScraperController, DEFAULT_OUTPUT_DIR
from app.province_mapping import get_province_output_file
from database_manager import DatabaseManager

# Initialize Database Manager (BigQuery only)
db_manager = DatabaseManager()

app = Flask(__name__, static_folder='static', template_folder='static')
app.config['SECRET_KEY'] = 'idealista-scraper-secret'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Allow embedding in iframes and add CORS headers for polling
@app.after_request
def after_request(response):
    response.headers.pop('X-Frame-Options', None)
    # Add CORS headers for cross-origin polling from dashboard
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/health', methods=['GET', 'OPTIONS'])
def health_check():
    """Simple health check endpoint for service readiness polling."""
    return jsonify({'status': 'ok'})

# Global scraper controller instance
scraper_controller: ScraperController | None = None
update_process = None
last_task_mode = None  # To track if update_process is 'update_urls' or 'enrichment'
session_start_time: float | None = None # Global start time for the current session/batch

# Scrape history storage (isolated per worker)
_wp = f"worker_{os.environ.get('SCRAPER_WORKER_ID', '')}_" if os.environ.get('SCRAPER_WORKER_ID') else ""
HISTORY_FILE = Path(__file__).parent.parent / f"{_wp}scrape_history.json"
scrape_history: list = []

# Caches for Excel file listing
EXCEL_FILES_GLOBAL_CACHE = {
    'timestamp': 0,
    'data': None
}
EXCEL_METADATA_CACHE = {}  # path -> {mtime, count}


def load_history():
    """Load scrape history from file."""
    global scrape_history
    if HISTORY_FILE.exists():
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Ensure data is a list, not a dict or other type
                if isinstance(data, list):
                    scrape_history = data
                else:
                    scrape_history = []
        except Exception:
            scrape_history = []
    return scrape_history


def save_history():
    """Save scrape history to file."""
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(scrape_history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving history: {e}")


def add_history_entry(seed_url: str, properties_count: int, category: str, output_file: str):
    """Add a new entry to scrape history."""
    entry = {
        'timestamp': datetime.now().isoformat(),
        'seed_url': seed_url,
        'properties_count': properties_count,
        'category': category,
        'output_file': output_file,
        'filename': os.path.basename(output_file) if output_file else None
    }
    scrape_history.insert(0, entry)  # Add to beginning (newest first)
    save_history()
    # Emit to connected clients
    socketio.emit('history_update', entry)
    return entry


# Load history on startup
load_history()


def emit_log(level: str, message: str):
    """Send log message to all connected clients."""
    socketio.emit('log', {'level': level, 'message': message})


def emit_property(data: dict):
    """Send scraped property data to all connected clients."""
    socketio.emit('property_scraped', data)


def emit_progress(data: dict):
    """Send progress update (pages/properties) to all connected clients."""
    socketio.emit('progress_update', data)


def emit_browser_closed():
    """Notify clients that browser was closed by user. Scraper is paused awaiting decision."""
    socketio.emit('browser_closed', {'message': 'Browser was closed. Resume or stop?'})


def emit_status(status: str, **kwargs):
    """Send status update to all connected clients."""
    # If we are in middle of a batch/periodic process, suppress individual 'completed' or 'stopped' statuses
    # arriving from sub-controllers, unless this event itself is the batch completion indicator.
    global periodic_process
    if periodic_process and periodic_process.poll() is None:
        if status in ('completed', 'stopped', 'idle') and kwargs.get('mode') != 'batch':
            # Suppress UI state change during batch sub-tasks
            return

    socketio.emit('status_change', {'status': status, **kwargs})
    
    # When completed OR stopped, add to history
    if status in ('completed', 'stopped') and scraper_controller:
        # Only add to history if there are scraped properties
        if scraper_controller.scraped_properties:
            category = scraper_controller._detected_sheet or 'unknown'
            add_history_entry(
                seed_url=scraper_controller.seed_url,
                properties_count=len(scraper_controller.scraped_properties),
                category=category,
                output_file=scraper_controller.output_file or ''
            )


@app.route('/')
def index():
    """Serve the main HTML interface with cache busting."""
    import time
    return render_template('index.html', cache_bust=int(time.time()))


@app.route('/api/config', methods=['GET'])
def get_config():
    """Get default configuration values."""
    return jsonify({
        'default_output_dir': DEFAULT_OUTPUT_DIR
    })


@app.route('/api/start', methods=['POST'])
def start_scraping():
    """Start a new scraping session."""
    global scraper_controller, session_start_time
    session_start_time = time.time()
    
    data = request.get_json()
    seed_url = data.get('seed_url', '').strip()
    mode = data.get('mode', 'stealth')  # 'stealth' or 'fast'
    dual_mode = data.get('dual_mode', False)
    output_dir = data.get('output_dir', '').strip() or DEFAULT_OUTPUT_DIR
    browser_engine = data.get('browser_engine', 'chromium')  # Multi-browser rotation
    smart_enrichment = data.get('smart_enrichment', False)  # Smart enrichment mode

    # Validate browser_engine
    if browser_engine not in ['chromium', 'firefox']:
        browser_engine = 'chromium'

    if not seed_url:
        return jsonify({'error': 'Seed URL is required'}), 400

    if not seed_url.startswith('http'):
        seed_url = 'https://' + seed_url

    if 'idealista.com' not in seed_url:
        return jsonify({'error': 'URL must be from idealista.com'}), 400

    # Stop any existing scraper
    if scraper_controller and scraper_controller.is_running:
        scraper_controller.stop()

    # For DUAL MODE: Calculate the second URL now
    dual_mode_url = None
    if dual_mode:
        if '/alquiler-viviendas/' in seed_url:
            dual_mode_url = seed_url.replace('/alquiler-viviendas/', '/venta-viviendas/')
        elif '/venta-viviendas/' in seed_url:
            dual_mode_url = seed_url.replace('/venta-viviendas/', '/alquiler-viviendas/')

    # Create controller with dual_mode_url if applicable
    scraper_controller = ScraperController(
        seed_url=seed_url,
        output_dir=output_dir,
        mode=mode,
        dual_mode_url=dual_mode_url,  # Pass second URL for same-browser execution
        browser_engine=browser_engine,  # Multi-browser rotation
        smart_enrichment=smart_enrichment,  # Smart enrichment mode
        on_log=emit_log,
        on_property=emit_property,
        on_status=emit_status,
        on_progress=emit_progress,
        on_browser_closed=emit_browser_closed,
        forced_target_file=data.get('target_file')  # Pass forced target file
    )
    
    # Start scraping in background thread
    def run_scraper():
        global scraper_controller
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            # Run scraper (handles both phases internally if dual_mode_url is set)
            loop.run_until_complete(scraper_controller.run())
        except Exception as e:
            emit_log("ERR", f"Scraper thread failed: {e}")
            if "CAPTCHA_BLOCK_DETECTED" in str(e):
                if scraper_controller: scraper_controller.status = "blocked"
                emit_status("blocked", message="Scraper blocked by CAPTCHA")
            else:
                if scraper_controller: scraper_controller.status = "error" 
                emit_status("error", message=str(e))
        finally:
            loop.close()
            # Ensure browser is closed
            if scraper_controller and scraper_controller.is_running:
                 try:
                     # This requires a slightly different way to call shutdown if loop is closed?
                     # ScraperController.stop() usually sets event.
                     pass
                 except: pass

    thread = threading.Thread(target=run_scraper, daemon=True)
    thread.start()
    
    return jsonify({'status': 'started', 'mode': mode, 'dual_mode': dual_mode})


@app.route('/api/debug/simulate_block', methods=['POST'])
def simulate_block():
    if scraper_controller:
        scraper_controller.status = "blocked"
        emit_status("blocked", message="Simulated CAPTCHA block")
    return jsonify({'status': 'blocked'})

@app.route('/api/set_mode', methods=['POST'])
def set_mode():
    """Update scraping mode dynamically."""
    data = request.get_json()
    mode = data.get('mode')
    
    if mode not in ['fast', 'stealth']:
        return jsonify({'error': 'Invalid mode'}), 400
    
    if scraper_controller:
        scraper_controller.set_mode(mode)
    
    # Also toggle flag for the standalone update_urls.py script
    try:
        update_script_dir = Path(__file__).parent.parent
        flag_file = update_script_dir / "update_stealth.flag"
        
        if mode == 'stealth':
            flag_file.touch()
        elif mode == 'fast':
            if flag_file.exists():
                flag_file.unlink()
    except Exception as e:
        print(f"Error toggling update mode flag: {e}")

    return jsonify({'status': 'mode_updated', 'mode': mode})





@app.route('/api/server/stop', methods=['POST'])
def stop_server():
    """Execute STOP_SILENT.bat to stop all services (no browser)."""
    try:
        base_dir = Path(__file__).parent.parent.parent
        bat_file = base_dir / 'scripts' / 'STOP_SILENT.bat'
        if bat_file.exists():
            subprocess.Popen(['cmd', '/c', str(bat_file)], cwd=str(base_dir),
                           creationflags=subprocess.CREATE_NO_WINDOW)
            return jsonify({'status': 'stopped', 'message': 'Server stopping...'})
        return jsonify({'error': 'STOP_SILENT.bat not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/server/restart', methods=['POST'])
def restart_server():
    """Execute RESTART_SILENT.bat to restart all services (no browser)."""
    try:
        base_dir = Path(__file__).parent.parent.parent
        bat_file = base_dir / 'scripts' / 'RESTART_SILENT.bat'
        if bat_file.exists():
            subprocess.Popen(['cmd', '/c', str(bat_file)], cwd=str(base_dir),
                           creationflags=subprocess.CREATE_NO_WINDOW)
            return jsonify({'status': 'restarting', 'message': 'Server restarting...'})
        return jsonify({'error': 'RESTART_SILENT.bat not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/pause', methods=['POST'])
def pause_scraping():
    """Pause the current scraping session (manual or batch)."""
    global scraper_controller, periodic_process
    
    paused = False
    # 1. Pause manual controller
    if scraper_controller and scraper_controller.is_running:
        scraper_controller.pause()
        paused = True
    
    # 2. Set flags for batch/periodic/update processes
    scraper_dir = Path(__file__).parent.parent
    try:
        (scraper_dir / "BATCH_PAUSE.flag").touch()
        (scraper_dir / "PERIODIC_PAUSE.flag").touch()
        (scraper_dir / "update_paused.flag").touch()
        
        if periodic_process and periodic_process.poll() is None:
            paused = True
        if update_process and update_process.poll() is None:
            paused = True
    except Exception as e:
        print(f"Error setting pause flags: {e}")
        
    if not paused:
        return jsonify({'error': 'No active scraping session'}), 400
    
    return jsonify({'status': 'paused'})


@app.route('/api/resume', methods=['POST'])
def resume_scraping():
    """Resume a paused scraping session (manual or batch)."""
    global scraper_controller, periodic_process
    
    resumed = False
    # 1. Resume manual controller
    if scraper_controller:
        scraper_controller.resume()
        resumed = True
    
    # 2. Remove flags for batch/periodic/update processes
    scraper_dir = Path(__file__).parent.parent
    try:
        batch_flag = scraper_dir / "BATCH_PAUSE.flag"
        periodic_flag = scraper_dir / "PERIODIC_PAUSE.flag"
        update_flag = scraper_dir / "update_paused.flag"
        
        if batch_flag.exists(): 
            batch_flag.unlink()
            resumed = True
        if periodic_flag.exists(): 
            periodic_flag.unlink()
            resumed = True
        if update_flag.exists():
            update_flag.unlink()
            resumed = True
    except Exception as e:
        print(f"Error removing pause flags: {e}")
        
    if not resumed:
        return jsonify({'error': 'No active scraping session to resume'}), 400
    
    return jsonify({'status': 'running'})


@app.route('/api/resume-state', methods=['GET'])
def get_resume_state():
    """Get saved resume state if available."""
    from app.scraper_wrapper import ScraperController
    state = ScraperController.load_state()
    if state:
        return jsonify({
            'has_state': True,
            'state': state
        })
    return jsonify({'has_state': False})


@app.route('/api/clear-state', methods=['POST'])
def clear_resume_state():
    """Clear saved resume state."""
    from app.scraper_wrapper import ScraperController
    ScraperController.clear_state()
    return jsonify({'status': 'cleared'})


@app.route('/api/stop', methods=['POST'])
def stop_scraping():
    """Stop the current scraping session (manual or batch) and export data."""
    global scraper_controller, periodic_process
    
    stopped = False
    
    # 1. Set stop flags first for scripts to exit gracefully
    scraper_dir = Path(__file__).parent.parent
    flags = ["BATCH_STOP.flag", "PERIODIC_STOP.flag", "update_stop.flag", "ENRICH_STOP.flag"]
    for f_name in flags:
        try: (scraper_dir / f_name).touch()
        except: pass

    # Synchronize: ensure batch_stop.flag is also set for any consumers checking it
    try: (scraper_dir / "batch_stop.flag").touch()
    except: pass

    # 2. Stop manual scraper controller
    if scraper_controller:
        scraper_controller.stop()
        stopped = True
        
    # Give background scripts a moment to see the flag
    time.sleep(0.5)

    # 3. Proactively terminate background processes
    if periodic_process and periodic_process.poll() is None:
        try:
            periodic_process.terminate()
            stopped = True
        except: pass
        
    if update_process and update_process.poll() is None:
        try:
            update_process.terminate()
            stopped = True
        except: pass

    if stopped:
        emit_status('stopped')
        return jsonify({'status': 'stopped'})
    else:
        return jsonify({'error': 'No active scraping session'}), 400
    
    return jsonify({'status': 'stopped'})


@app.route('/api/status', methods=['GET'])
def get_status():
    """Get current scraper status (manual, batch, or periodic)."""
    global scraper_controller, periodic_process
    
    status = 'idle'
    mode = 'fast'
    properties_count = 0
    output_file = None
    
    if scraper_controller:
        status = scraper_controller.status
        properties_count = len(scraper_controller.scraped_properties)
        output_file = scraper_controller.output_file
        mode = getattr(scraper_controller, 'mode', 'fast')
        
    # Check if batch runner (periodic_process) is active
    if periodic_process and periodic_process.poll() is None:
        # If batch is running but individual scraper has an error/stopped, 
        # report 'running' to keep UI locked until batch actually finishes.
        if status in ('idle', 'completed', 'stopped', 'error'):
            status = 'running'
        mode = 'batch'
        
        # Check for pause flags to report overall status correctly
        scraper_dir = Path(__file__).parent.parent
        if (scraper_dir / "BATCH_PAUSE.flag").exists() or (scraper_dir / "PERIODIC_PAUSE.flag").exists():
            status = 'paused'

    # Check for running URL update process
    global update_process
    if update_process and update_process.poll() is None:
        if status in ('idle', 'completed', 'stopped', 'error'):
            status = 'running'
        mode = 'update_urls'
        # Check for URL update specifically being paused
        if (Path(__file__).parent.parent / "update_paused.flag").exists():
            status = 'paused'
            
    current_page = scraper_controller.current_page if scraper_controller else 0
    total_pages = scraper_controller.total_pages_expected if scraper_controller else 0
    total_properties = scraper_controller.total_properties_expected if scraper_controller else 0
    
    # Use global session_start_time as priority for batches, 
    # fallback to controller's start_time for manual scrapes
    start_time_resp = session_start_time
    if scraper_controller and not start_time_resp:
        start_time_resp = scraper_controller.start_time

    return jsonify({
        'status': status,
        'internal_status': scraper_controller.status if scraper_controller else 'idle',
        'properties_count': properties_count,
        'current_page': current_page,
        'total_pages': total_pages,
        'total_properties': total_properties,
        'start_time': start_time_resp,
        'output_file': output_file,
        'mode': mode,
        'task_mode': last_task_mode
    })


@app.route('/api/batch/status', methods=['GET'])
def get_batch_status():
    """Detailed status for the batch enrichment process."""
    global update_process, last_task_mode
    
    is_running = update_process and update_process.poll() is None and last_task_mode == 'enrichment'
    
    # We can't easily get current_idx and total without reading logs or a shared state,
    # but for now reporting is_running is enough to unlock the UI.
    return jsonify({
        'is_running': is_running,
        'task': 'enrichment' if last_task_mode == 'enrichment' else 'none'
    })


# Periodic Low-Cost Scraper Process
periodic_process = None
periodic_thread = None

def periodic_log_monitor(process):
    """Refined monitor to stream logs via specific socket event."""
    try:
        # Read stdout line by line
        for line in iter(process.stdout.readline, ''):
            decoded = line.strip()
            if decoded:
                socketio.emit('periodic_log', {'message': decoded})
                
                # Try to parse structure for table updates (Simple parsing for now)
                if "[STATUS]" in decoded:
                    status = decoded.split("[STATUS]")[1].strip().lower()
                    emit_status(status)
                elif "[OK] Scrape completed for" in decoded:
                    prov = decoded.split("for")[-1].strip().replace(".", "")
                    socketio.emit('periodic_table_update', {'province': prov, 'status': 'Completado'})
                elif "Processing:" in decoded:
                    prov = decoded.split("Processing:")[-1].strip()
                    socketio.emit('periodic_table_update', {'province': prov, 'status': 'Procesando...'})
                
                # ALSO emit to main console log for visibility
                level = 'INFO'
                msg = decoded
                if '[ERR]' in decoded: 
                    level = 'ERR'
                    msg = decoded.replace('[ERR]', '').strip()
                elif '[WARN]' in decoded: 
                    level = 'WARN'
                    msg = decoded.replace('[WARN]', '').strip()
                elif '[OK]' in decoded: 
                    level = 'OK'
                    msg = decoded.replace('[OK]', '').strip()
                
                emit_log(level, msg)

        process.stdout.close()
        process.wait()
        
        # Emit completion status with more detail
        if process.returncode == 0:
            emit_log('OK', "✅ Proceso background finalizado correctamente.")
            emit_status('completed', mode='batch', message="Proceso batch finalizado correctamente")
        else:
            emit_log('ERR', f"❌ Proceso background falló (Código {process.returncode}).")
            emit_status('error', mode='batch', message=f"Proceso finalizado con errores (Código {process.returncode})")
            
    except Exception as e:
        print(f"Monitor error: {e}")
        emit_log('ERR', f"❌ Error en monitor de logs: {str(e)}")
        emit_status('error', message=f"Error en monitor: {str(e)}")
    finally:
        # CRITICAL: Clear global process state to allow starting new batches/periodic scans
        global periodic_process, periodic_thread
        periodic_process = None
        periodic_thread = None
        
        # Cleanup flags
        scraper_dir = Path(__file__).parent.parent
        for f_name in ["BATCH_STOP.flag", "PERIODIC_STOP.flag", "BATCH_PAUSE.flag", "PERIODIC_PAUSE.flag"]:
            try: (scraper_dir / f_name).unlink()
            except: pass

@app.route('/api/periodic-lowcost/start', methods=['POST'])
def start_periodic_lowcost():
    """Launch the periodic low-cost scraper in a background process."""
    global periodic_process, periodic_thread, session_start_time
    session_start_time = time.time()
    
    if periodic_process and periodic_process.poll() is None:
        return jsonify({'error': 'Periodic scan already running'}), 400
    
    data = request.get_json() or {}
    operation = data.get('operation', 'sale') # Default to sale as requested by context
    
    script_path = Path(__file__).parent.parent.parent / "scripts" / "run_periodic_low_cost.py"
    scraper_dir = script_path.parent.parent / "scraper"
    
    # Cleanup flags
    stop_flag = scraper_dir / "PERIODIC_STOP.flag"
    pause_flag = scraper_dir / "PERIODIC_PAUSE.flag"
    if stop_flag.exists(): os.remove(stop_flag)
    if pause_flag.exists(): os.remove(pause_flag)
    
    if not script_path.exists():
        return jsonify({'error': f'Script not found: {script_path}'}), 500
    
    # Launch in background with PIPE for logging
    cmd = [sys.executable, str(script_path), "--operation", operation]
    if data.get('use_vpn'):
        cmd.append("--nordvpn")

    periodic_process = subprocess.Popen(
        cmd,
        cwd=str(script_path.parent.parent),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT, # Merge stderr to stdout
        text=True,
        encoding='utf-8',
        errors='replace',
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
    )
    
    # Start monitor thread
    periodic_thread = threading.Thread(target=periodic_log_monitor, args=(periodic_process,), daemon=True)
    periodic_thread.start()
    
    return jsonify({'status': 'started', 'pid': periodic_process.pid})


@app.route('/api/periodic-lowcost/stop', methods=['POST'])
def stop_periodic_lowcost():
    scraper_dir = Path(__file__).parent.parent.parent / "scraper"
    flag = scraper_dir / "PERIODIC_STOP.flag"
    with open(flag, 'w') as f: f.write("STOP")
    return jsonify({'status': 'stopping'})

@app.route('/api/periodic-lowcost/pause', methods=['POST'])
def pause_periodic_lowcost():
    scraper_dir = Path(__file__).parent.parent.parent / "scraper"
    flag = scraper_dir / "PERIODIC_PAUSE.flag"
    with open(flag, 'w') as f: f.write("PAUSE")
    return jsonify({'status': 'paused'})

@app.route('/api/periodic-lowcost/resume', methods=['POST'])
def resume_periodic_lowcost():
    scraper_dir = Path(__file__).parent.parent.parent / "scraper"
    flag = scraper_dir / "PERIODIC_PAUSE.flag"
    if flag.exists(): os.remove(flag)
    return jsonify({'status': 'resumed'})

@app.route('/api/periodic-lowcost/status', methods=['GET'])
def get_periodic_status():
    """Get status of the periodic low-cost scraper."""
    global periodic_process
    
    status = 'not_started'
    if periodic_process:
        poll = periodic_process.poll()
        if poll is None:
            status = 'running'
            # Check pause
            scraper_dir = Path(__file__).parent.parent.parent / "scraper"
            if (scraper_dir / "PERIODIC_PAUSE.flag").exists():
                status = 'paused'
        else:
            status = 'completed'
            
    return jsonify({'status': status})


@app.route('/api/download', methods=['GET'])
def download_file():
    """Download the generated Excel file."""
    global scraper_controller
    
    if not scraper_controller or not scraper_controller.output_file:
        return jsonify({'error': 'No file available for download'}), 404
    
    file_path = scraper_controller.output_file
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(
        file_path,
        as_attachment=True,
        download_name=os.path.basename(file_path)
    )


@app.route('/api/properties', methods=['GET'])
def get_properties():
    """Get all scraped properties."""
    global scraper_controller
    
    if not scraper_controller:
        return jsonify({'properties': []})
    
    return jsonify({'properties': scraper_controller.scraped_properties})


@app.route('/api/history', methods=['GET'])
def get_history():
    """Get scrape history."""
    return jsonify({'history': scrape_history})


@app.route('/api/history/clear', methods=['POST'])
def clear_history():
    """Clear scrape history."""
    global scrape_history
    scrape_history = []
    save_history()
    return jsonify({'status': 'cleared'})


@app.route('/api/excel-files', methods=['GET'])
def get_excel_files():
    """Get list of Excel files in the output directory and project directories."""
    global EXCEL_FILES_GLOBAL_CACHE, EXCEL_METADATA_CACHE
    import pandas as pd
    
    # Check TTL cache (5 seconds)
    if time.time() - EXCEL_FILES_GLOBAL_CACHE['timestamp'] < 5:
        if EXCEL_FILES_GLOBAL_CACHE['data']:
            return jsonify(EXCEL_FILES_GLOBAL_CACHE['data'])
            
    files = []
    
    # Search in multiple directories
    search_dirs = [
        DEFAULT_OUTPUT_DIR,  # Primary output directory
        Path(__file__).parent.parent,  # Scraper root directory
        Path(DEFAULT_OUTPUT_DIR).parent if DEFAULT_OUTPUT_DIR else None,  # Parent of output
    ]
    
    for search_dir in search_dirs:
        if not search_dir:
            continue
        search_path = Path(search_dir)
        if search_path.exists():
            for f in search_path.glob('*.xlsx'):
                if f.is_file():
                    path_str = str(f.resolve())
                    mtime = f.stat().st_mtime
                    
                    # Check metadata cache
                    if path_str in EXCEL_METADATA_CACHE and EXCEL_METADATA_CACHE[path_str]['mtime'] == mtime:
                        count = EXCEL_METADATA_CACHE[path_str]['count']
                    else:
                        # count rows (this is the slow part)
                        count = 0
                        try:
                            # Optimization: Use openpyxl directly to count rows if possible, 
                            # or just read the first column of each sheet.
                            dfs = pd.read_excel(f, sheet_name=None, usecols=[0], engine='openpyxl')
                            count = sum(len(df) for df in dfs.values())
                            
                            # Store in metadata cache
                            EXCEL_METADATA_CACHE[path_str] = {
                                'mtime': mtime,
                                'count': count
                            }
                        except Exception as e:
                            print(f"Error counting rows in {f.name}: {e}")
                            pass
                        
                    files.append({
                        'name': f.name,
                        'path': path_str,
                        'count': count,
                        'mtime': mtime
                    })
    
    # Deduplicate by path
    seen_paths = set()
    unique_files = []
    for f in files:
        if f['path'] not in seen_paths:
            seen_paths.add(f['path'])
            unique_files.append(f)
            
    # Sort by mtime (newest first)
    unique_files.sort(key=lambda x: x.get('mtime', 0), reverse=True)
    
    result = {'files': unique_files}
    
    # Update global cache
    EXCEL_FILES_GLOBAL_CACHE['timestamp'] = time.time()
    EXCEL_FILES_GLOBAL_CACHE['data'] = result
    
    return jsonify(result)


@app.route('/api/provinces-list', methods=['GET'])
def get_provinces_list():
    """Return list of Spanish provinces with verified venta and alquiler URLs, PLUS nested zones."""
    try:
        # 1. Load basic provinces (ID, Name, URLs)
        json_path = Path(__file__).parent.parent / "low_cost_provinces.json"
        if not json_path.exists():
            return jsonify({'error': 'Provinces file not found'}), 404
            
        with open(json_path, 'r', encoding='utf-8') as f:
            provinces = json.load(f)
            
        # 2. Load extracted zones mapping
        zones_path = Path(__file__).parent.parent / "province_zones_complete.json"
        zones_map = {}
        if zones_path.exists():
            with open(zones_path, 'r', encoding='utf-8') as f:
                zones_map = json.load(f)

        # 3. Load subzones mapping
        subzones_path = Path(__file__).parent.parent / "documentation" / "subzones_complete.json"
        subzones_map = {}
        if subzones_path.exists():
            with open(subzones_path, 'r', encoding='utf-8') as f:
                subzones_map = json.load(f)

        # 4. Merge zones + subzones into provinces list
        for p in provinces:
            # Match by Name (e.g. "A Coruña")
            p_name = p.get('name')
            if p_name and p_name in zones_map:
                zones = zones_map[p_name].get('zones', [])
                prov_subzones = subzones_map.get(p_name, {})
                for zone in zones:
                    zone_name = zone.get('name', '')
                    zone_data = prov_subzones.get(zone_name, {})
                    zone['subzones'] = zone_data.get('subzones', [])
                p['zones'] = zones
            else:
                p['zones'] = []

        # Return the merged data
        return jsonify({'provinces': provinces})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def expand_batch_urls(urls):
    """
    Disabled expansion to allow direct use of province URLs from frontend.
    """
    return urls



@app.route('/api/start-batch', methods=['POST'])
def start_batch_scraping():
    """Start a batch scraping process for a list of URLs."""
    global periodic_process, periodic_thread, session_start_time
    session_start_time = time.time()
    
    data = request.json
    print(f"DEBUG: start_batch received: {data}")
    urls = data.get('urls', [])
    mode = data.get('mode', 'fast')
    
    if not urls:
        return jsonify({'error': 'No URLs provided'}), 400
        
    # Filter out None/Empty
    urls = [u for u in urls if u and isinstance(u, str) and u.strip()]
    
    if not urls:
        return jsonify({'error': 'No valid URLs after filtering'}), 400
        
    if periodic_process and periodic_process.poll() is None:
        return jsonify({'error': 'A batch process is already running'}), 400
        
    
    # Expand provincial URLs if needed (default to True for backward compat, but app.js can disable)
    should_expand = data.get('expand', True)
    if should_expand:
        original_count = len(urls)
        urls = expand_batch_urls(urls)
        if len(urls) > original_count:
            print(f"Batch expanded from {original_count} to {len(urls)} URLs")
    else:
        print("Batch expansion disabled by client. Applying Intelligent Zone Segmentation instead if needed...")
        
    # === INTELLIGENT ZONE TARGETING ===
    # Check if we should expand provinces exceeding 2000 properties into zones
    try:
        import sqlite3
        db_path = Path(__file__).parent.parent.parent / "real_estate.db"
        zones_file = Path(__file__).parent.parent / "province_zones_complete.json"
        
        if db_path.exists() and zones_file.exists():
            with open(zones_file, 'r', encoding='utf-8') as zf:
                zones_map = json.load(zf)
            
            conn = sqlite3.connect(db_path)
            c = conn.cursor()
            
            def local_extract_province(url):
                try:
                    import urllib.parse
                    import re
                    path = urllib.parse.urlsplit(url).path
                    parts = [p for p in path.split('/') if p]
                    if 'alquiler-viviendas' in parts:
                        idx = parts.index('alquiler-viviendas')
                    elif 'venta-viviendas' in parts:
                        idx = parts.index('venta-viviendas')
                    else:
                        return "Desconocida"
                    if len(parts) > idx + 1:
                        raw = parts[idx + 1]
                        return raw.replace('-', ' ').title()
                except: pass
                return "Desconocida"
            
            expanded_urls = []
            
            for url in urls:
                try:
                    prov_name = local_extract_province(url)
                    op_type = 'venta' if '/venta-' in url else 'alquiler'
                    
                    if prov_name:
                        c.execute('''
                            SELECT total_properties FROM inventory_trends 
                            WHERE province = ? AND operation = ? AND zone = '(Toda la provincia)'
                            ORDER BY date_record DESC LIMIT 1
                        ''', (prov_name, op_type))
                        row = c.fetchone()
                        
                        if row and row[0] > 2000:
                            print(f"[SMART SEGMENTATION] {prov_name} ({op_type}) has {row[0]} properties (>2000). Segmenting into zones...")
                            # Find province in zones_map
                            prov_zones = []
                            for p_key, p_data in zones_map.items():
                                if p_key.lower() == prov_name.lower() or p_data.get('name', '').lower() == prov_name.lower():
                                    prov_zones = p_data.get('zones', [])
                                    break
                                    
                            if prov_zones:
                                for zone in prov_zones:
                                    z_href = zone.get('href')
                                    if z_href:
                                        if op_type == 'alquiler' and '/venta-viviendas/' in z_href:
                                            z_href = z_href.replace('/venta-viviendas/', '/alquiler-viviendas/')
                                        elif op_type == 'venta' and '/alquiler-viviendas/' in z_href:
                                            z_href = z_href.replace('/alquiler-viviendas/', '/venta-viviendas/')
                                            
                                        # Also respect price limit if present in original URL
                                        if "con-precio-hasta_" in url:
                                            price_part = [p for p in url.split("/") if "con-precio-" in p]
                                            if price_part:
                                                if not z_href.endswith('/'): z_href += '/'
                                                z_href += price_part[0] + '/'
                                                
                                        full_url = f"https://www.idealista.com{z_href}"
                                        expanded_urls.append(full_url)
                                continue  # Skip adding the raw province url
                                
                    expanded_urls.append(url)
                except Exception as ex:
                    print(f"Error checking zone segmentation for {url}: {ex}")
                    expanded_urls.append(url)
            
            urls = expanded_urls
            conn.close()
    except Exception as e:
        print(f"Intelligent Zone Targeting skipped due to inner exception: {e}")
        
    # Write queue to file
    queue_file = Path(__file__).parent.parent / "batch_queue.json"
    smart_enrichment = data.get('smart_enrichment', False)
    target_file = data.get('target_file')
    province_name = data.get('province_name')
    operation_type = data.get('operation_type')

    # Auto-resolve target_file if missing but province/operation are provided
    if not target_file and province_name and operation_type:
        target_file = get_province_output_file(province_name, operation_type)
        if target_file:
            print(f"Auto-resolved target file for {province_name} ({operation_type}): {target_file}")

    with open(queue_file, 'w', encoding='utf-8') as f:
        json.dump({
            'urls': urls,
            'mode': mode,
            'smart_enrichment': smart_enrichment,
            'target_file': target_file
        }, f)
        
    # Spawn runner
    script_path = Path(__file__).parent.parent.parent / "scripts" / "run_batch.py"
    if not script_path.exists():
        # Fallback create if not exists (we will create it next)
        pass 
        
    scraper_dir = Path(__file__).parent.parent
    
    # Reset flags
    stop_flag = scraper_dir / "BATCH_STOP.flag"
    pause_flag = scraper_dir / "BATCH_PAUSE.flag"
    if stop_flag.exists(): os.remove(stop_flag)
    if pause_flag.exists(): os.remove(pause_flag)

    periodic_process = subprocess.Popen(
        [sys.executable, str(script_path)],
        cwd=str(scraper_dir),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True, # Critical for line buffering
        encoding='utf-8',
        errors='replace',
        bufsize=1, # Line buffered
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
    )
    
    # Reuse periodic log monitor (renamed mentally to batch monitor)
    periodic_thread = threading.Thread(target=periodic_log_monitor, args=(periodic_process,), daemon=True)
    periodic_thread.start()
    
    # Emit initial running status for batch
    emit_status('running', mode='batch')
    
    return jsonify({'status': 'started', 'pid': periodic_process.pid, 'count': len(urls)})


@app.route('/api/batch/stop', methods=['POST'])
def stop_batch_scraping():
    """Stop the batch scraping process (Periodic or Enrichment)."""
    scraper_dir = Path(__file__).parent.parent
    
    # Touch ALL relevant stop flags
    (scraper_dir / "BATCH_STOP.flag").touch()
    (scraper_dir / "batch_stop.flag").touch()
    (scraper_dir / "ENRICH_STOP.flag").touch()
    (scraper_dir / "update_stop.flag").touch()
    
    # Stop the active scraper controller
    global scraper_controller
    if scraper_controller:
        scraper_controller.stop()
    
    # Terminate periodic process (if running)
    global periodic_process
    if periodic_process and periodic_process.poll() is None:
        try:
            periodic_process.terminate()
            print("[server] Terminated periodic/batch process via batch/stop")
        except: pass

    # Terminate update/enrichment process (if running)
    global update_process
    if update_process and update_process.poll() is None:
        try:
            update_process.terminate()
            print("[server] Terminated update/enrichment process via batch/stop")
        except: pass
    return jsonify({'status': 'stopping'})


@app.route('/api/batch/pause', methods=['POST'])
def pause_batch_scraping():
    """Pause the batch scraping process."""
    scraper_dir = Path(__file__).parent.parent
    flag = scraper_dir / "BATCH_PAUSE.flag"
    with open(flag, 'w') as f: f.write("PAUSE")
    
    # Also pause active controller if any
    if scraper_controller:
        scraper_controller.pause()
        
    return jsonify({'status': 'paused'})


@app.route('/api/batch/resume', methods=['POST'])
def resume_batch_scraping():
    """Resume the batch scraping process."""
    scraper_dir = Path(__file__).parent.parent
    flag = scraper_dir / "BATCH_PAUSE.flag"
    if flag.exists(): os.remove(flag)
    
    # Also resume active controller if any
    if scraper_controller:
        scraper_controller.resume()
        
    return jsonify({'status': 'resumed'})



@app.route('/api/excel-worksheets', methods=['GET'])
def get_excel_worksheets():
    """Get list of worksheet names for a given Excel file."""
    import pandas as pd
    
    file_path = request.args.get('file', '').strip()
    
    if not file_path:
        return jsonify({'error': 'File path is required'}), 400
    
    if not os.path.exists(file_path):
        return jsonify({'error': f'File not found: {file_path}'}), 404
    
    try:
        # Read just sheet names without loading data
        xl = pd.ExcelFile(file_path)
        sheets = xl.sheet_names
        return jsonify({'sheets': sheets})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/update-urls', methods=['POST'])
def update_urls():
    """Start URL update process for an existing Excel file."""
    import subprocess
    import json as json_module
    
    data = request.get_json()
    excel_file = data.get('excel_file', '').strip()
    sheets = data.get('sheets', [])  # List of sheet names to process
    resume = data.get('resume', False)
    
    if not excel_file:
        return jsonify({'error': 'Excel file path is required'}), 400
    
    if not os.path.exists(excel_file):
        return jsonify({'error': f'File not found: {excel_file}'}), 404
    
    # Path to the update_urls.py script (local copy in scraper workspace)
    update_script = Path(__file__).parent.parent / 'update_urls.py'
    
    if not update_script.exists():
        return jsonify({'error': 'Update script not found'}), 500
    
    emit_log('INFO', f'Starting URL update for: {os.path.basename(excel_file)}')
    
    # Run the update script in background with file path argument
    def run_update():
        global update_process
        import subprocess
        try:
            # Clean flag before start
            flag_file = update_script.parent / "update_paused.flag"
            if flag_file.exists():
                flag_file.unlink()

            # Run script with the Excel file path and sheets as arguments
            sheets_json = json_module.dumps(sheets) if sheets else '[]'
            
            # Use currentMode from frontend (relayed via server state or global)
            # Since server.py doesn't track currentMode directly, we use the flag to decide initial mode
            # but we can pass it explicitly if we know it. For now, rely on consistency.
            mode_arg = 'stealth' if (update_script.parent / "update_stealth.flag").exists() else 'fast'

            cmd = ['python', '-u', str(update_script), excel_file, '--sheets', sheets_json, '--mode', mode_arg]
            if resume:
                cmd.append('--resume')

            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True, # text=True is important for readline
                bufsize=1, # Line buffered
                cwd=str(update_script.parent)
            )
            update_process = process

            
            # Stream output line by line to WebSocket
            for line in iter(process.stdout.readline, ''):
                line = line.strip()
                if not line:
                    continue
                # Parse log level from line
                # Parse log level from line
                if '[STATUS]' in line:
                    status = line.split('[STATUS]')[1].strip().lower() # e.g. "paused"
                    emit_status(status)
                elif '[ERR]' in line or 'ERROR' in line:
                    emit_log('ERR', line.replace('[ERR]', '').replace('ERROR:', '').strip())
                elif '[WARN]' in line:
                    emit_log('WARN', line.replace('[WARN]', '').strip())
                elif '[OK]' in line or 'SUCCESS' in line:
                    emit_log('OK', line.replace('[OK]', '').replace('SUCCESS:', '').strip())
                elif '[INFO]' in line:
                    emit_log('INFO', line.replace('[INFO]', '').strip())
                else:
                    emit_log('INFO', line)
            
            process.wait()
            
            # Check if stopped manually (global var cleared)
            was_stopped = (update_process is None)
            update_process = None
            
            # Clean up flags
            if flag_file.exists():
                flag_file.unlink()
            stop_flag = update_script.parent / "update_stop.flag"
            if stop_flag.exists():
                stop_flag.unlink()
            
            if was_stopped:
                 emit_log('INFO', 'Update stopped by user.')
                 emit_status('stopped', message='Update stopped')
            elif process.returncode == 0:
                emit_log('OK', 'URL update completed successfully!')
                emit_status('completed', message='URL update finished')
            else:
                emit_log('ERR', f'Update failed with code {process.returncode}')
                emit_status('error', message='URL update failed')
                
        except Exception as e:
            emit_log('ERR', f'Error running update: {str(e)}')
            emit_status('error', message=str(e))
    
    thread = threading.Thread(target=run_update, daemon=True)
    thread.start()
    
    # Emit initial running status
    emit_status('running', mode='update_urls')
    
    return jsonify({'status': 'started', 'file': excel_file})


@app.route('/api/update/check-state', methods=['POST'])
def check_update_state():
    """Check if there is a resumable state for the given file."""
    import json as json_module
    
    data = request.get_json()
    excel_file = data.get('excel_file', '').strip()
    
    if not excel_file:
        return jsonify({'error': 'Excel file path is required'}), 400
        
    update_script = Path(__file__).parent.parent / "update_urls.py"
    journal_file = update_script.parent / "update_progress.jsonl"
    
    if not journal_file.exists():
        return jsonify({'can_resume': False})
        
    try:
        # Check if journal matches this file by reading the first line
        with open(journal_file, 'r', encoding='utf-8') as f:
            first_line = f.readline()
            if not first_line:
                return jsonify({'can_resume': False})
                
            entry = json_module.loads(first_line)
            if entry.get('full_path') == excel_file:
                # Count lines to determine progress
                # Reset file pointer to count all
                f.seek(0)
                count = sum(1 for _ in f)
                
                # We don't know total here easily unless we open Excel, 
                # but we can return the count of processed items.
                # Ideally we should cache the total?
                # For now let's just return the count and client can disable "Resume" if data is weird.
                return jsonify({
                    'can_resume': True,
                    'current_index': count,
                    'total': '?' # Client will show "Reanudar (X finished)"
                })
    except:
        pass
        
    return jsonify({'can_resume': False})



@app.route('/api/update/pause', methods=['POST'])
def pause_update():
    """Pause the update process."""
    update_script = Path(__file__).parent.parent / "update_urls.py"
    flag_file = update_script.parent / "update_paused.flag"
    flag_file.touch()
    return jsonify({'status': 'paused'})

@app.route('/api/update/resume', methods=['POST'])
def resume_update():
    """Resume the update process."""
    update_script = Path(__file__).parent.parent / "update_urls.py"
    flag_file = update_script.parent / "update_paused.flag"
    if flag_file.exists():
        flag_file.unlink()
    return jsonify({'status': 'resumed'})

@app.route('/api/update/stop', methods=['POST'])
def stop_update_process():
    """Stop the update process."""
    global update_process, scraper_controller
    update_script = Path(__file__).parent.parent / "update_urls.py"
    scraper_dir = Path(__file__).parent.parent

    # 1. Signal internal controller
    if scraper_controller:
        scraper_controller.stop()
    
    # 2. Set ALL stop flags (Update mode often triggers scraper runs)
    flags = ["update_stop.flag", "BATCH_STOP.flag", "batch_stop.flag"]
    for f_name in flags:
        try: (scraper_dir / f_name).touch()
        except: pass
    
    try:
        if update_process:
            try:
                time.sleep(0.5)
                update_process.terminate()
            except:
                pass
            update_process = None
            
        # Clean pause flag
        flag_file = update_script.parent / "update_paused.flag"
        if flag_file.exists():
            flag_file.unlink()
            
        return jsonify({'status': 'stopped'})
    except Exception as e:
        print(f"Error stopping update: {e}")
        return jsonify({'error': str(e)}), 500



# /api/import-api removed.


@socketio.on('progress')
def handle_progress(data):
    """Forward progress events from update_urls.py to UI."""
    socketio.emit('progress_update', data)

@socketio.on('log_message')
def handle_relay_log(data):
    """Relay log messages from background scripts to UI."""
    socketio.emit('log', data)

@socketio.on('property_scraped')
def handle_relay_property(data):
    """Relay property data from background scripts to UI."""
    socketio.emit('property_scraped', data)


def run_server(host='127.0.0.1', port=None):
    """Run the Flask-SocketIO server."""
    if port is None:
        port = int(os.environ.get("SCRAPER_PORT", 5003))
    print(f"Starting Scraper Server on port {port}...")
    socketio.run(app, host=host, port=port, debug=False, allow_unsafe_werkzeug=True)



# =============================================================================
# API & DATABASE DASHBOARD ENDPOINTS
# =============================================================================

@app.route('/api/salidas-files', methods=['GET'])
def get_salidas_files():
    """Optimized file listing for scraper/salidas using scandir for maximum performance."""
    try:
        limit = request.args.get('limit', default=200, type=int)

        # Resolve scraper/salidas relative to this server file
        current_dir = Path(__file__).parent.parent
        salidas_dir = (current_dir / "salidas").resolve()

        if not salidas_dir.exists():
            return jsonify({'files': []})

        files = []
        # scandir is significantly faster for directories with many files
        with os.scandir(salidas_dir) as it:
            for entry in it:
                if entry.is_file() and entry.name.endswith('.xlsx') and not entry.name.startswith('~$') and not entry.name.startswith('.'):
                    files.append({
                        'name': entry.name,
                        'path': entry.path,
                        'mtime': entry.stat().st_mtime
                    })

        # Sort by modification time (newest first)
        files.sort(key=lambda x: x['mtime'], reverse=True)

        # Apply limit
        files = files[:limit]

        # Count total rows across all sheets per file
        try:
            from openpyxl import load_workbook
            for f in files:
                try:
                    wb = load_workbook(f['path'], read_only=True, data_only=True)
                    total = 0
                    for ws in wb.worksheets:
                        rows = ws.max_row
                        if rows and rows > 1:
                            total += rows - 1  # subtract header
                    wb.close()
                    f['count'] = total
                except Exception:
                    f['count'] = None
        except ImportError:
            pass

        return jsonify({'files': files})
    except Exception as e:
        print(f"Error in get_salidas_files: {e}")
        return jsonify({'error': str(e), 'files': []}), 500

# /api/batch-scan removed.


@app.route('/api/batch/start', methods=['POST'])
def start_batch_enrichment():
    """Start batch enrichment for one or more Excel files."""
    global update_process, last_task_mode, session_start_time
    session_start_time = time.time()
    if update_process and update_process.poll() is None:
        return jsonify({'status': 'error', 'error': 'A task is already running. Please wait.'}), 409

    data = request.json or {}
    file_paths = data.get('files', [])

    if not file_paths:
        return jsonify({'error': 'No files provided'}), 400

    # Validate all files exist
    for fp in file_paths:
        if not os.path.exists(fp):
            return jsonify({'error': f'File not found: {fp}'}), 404

    # Build command: run enrich_worker.py for each file sequentially
    # The script accepts --input as a glob, so we pass each file individually
    # We use a wrapper approach: call the script once per file
    script_path = (Path(__file__).parent.parent.parent / "scripts" / "enrich_worker.py").resolve()

    if not script_path.exists():
        return jsonify({'error': 'enrich_worker.py script not found'}), 500

    # For batch: pass all files as separate --input args via a small inline script
    # OR: we can call enrich_worker.py once per file sequentially
    # Simplest approach: create a temporary batch command that processes all files
    if len(file_paths) == 1:
        cmd = [sys.executable, str(script_path), "--input", file_paths[0]]
    else:
        # For multiple files, we create a small inline Python wrapper
        # that calls enrich_worker.py for each file
        inline_script = "; ".join([
            "import subprocess, sys",
            f"files = {file_paths!r}",
            f"script = r'{script_path}'",
            "for f in files:",
            "    print(f'\\n=== Processing: {f} ===', flush=True)",
            "    rc = subprocess.call([sys.executable, script, '--input', f])",
            "    if rc != 0: print(f'WARNING: {f} exited with code {rc}', flush=True)"
        ])
        cmd = [sys.executable, "-c", inline_script]

    return start_background_task(cmd, f"Batch Enrichment ({len(file_paths)} files)", task_mode='enrichment')

@app.route('/api/enrich', methods=['POST'])
def run_enrichment():
    """Run enrichment worker script."""
    global update_process
    if update_process and update_process.poll() is None:
        return jsonify({'status': 'error', 'message': 'A task is already running. Please wait.'}), 409
    
    data = request.json or {}
    operation = data.get('operation', 'rent')
    file_path = data.get('file_path')
    
    script_path = (Path(__file__).parent.parent.parent / "scripts" / "enrich_worker.py").resolve()
    
    if file_path:
        # Use specific file (from picker)
        input_pattern = file_path
    else:
        # Fallback to operation pattern
        input_pattern = f"scraper/salidas/*_{operation}_*.xlsx"
    
    cmd = [sys.executable, str(script_path), "--input", input_pattern, "--max-price", "300000"]
    return start_background_task(cmd, f"Enrichment ({operation.upper()})", task_mode='enrichment')

# Supabase endpoints removed


def start_background_task(cmd, task_name, cwd=None, task_mode='update_urls'):
    """Helper to start a background process and stream output to frontend."""
    global update_process, last_task_mode
    last_task_mode = task_mode
    
    # Clear ALL stop flags before launching to prevent stale flags from killing new processes
    scraper_dir = Path(__file__).parent.parent
    for f_name in ["BATCH_STOP.flag", "PERIODIC_STOP.flag", "update_stop.flag", "ENRICH_STOP.flag"]:
        flag_path = scraper_dir / f_name
        if flag_path.exists():
            try:
                flag_path.unlink()
                print(f"[server] Cleared stale flag: {f_name}")
            except Exception:
                pass
    
    def run_and_stream():
        global update_process
        emit_log("INFO", f"Starting task: {task_name}")
        emit_log("INFO", f"Command: {' '.join(cmd)}")
        
        try:
            # Use project root as default CWD if not specified
            working_dir = cwd if cwd else str(Path(__file__).parent.parent.parent)
            
            update_process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                errors='replace',
                bufsize=1,
                cwd=working_dir,
                env={**os.environ, "PYTHONUNBUFFERED": "1"} # Force unbuffered output
            )
            
            emit_status("running", mode="batch")
            
            for line in iter(update_process.stdout.readline, ''):
                if line:
                    emit_log("INFO", line.strip())
            
            update_process.wait()
            rc = update_process.returncode
            
            scraper_dir = Path(__file__).parent.parent
            stop_flag_exists = any((scraper_dir / f).exists() for f in ["BATCH_STOP.flag", "PERIODIC_STOP.flag", "update_stop.flag", "ENRICH_STOP.flag"])

            if rc == 0:
                emit_log("OK", f"Task '{task_name}' completed successfully.")
                emit_status("completed")
            elif stop_flag_exists:
                emit_log("INFO", f"Task '{task_name}' was stopped by user.")
                emit_status("stopped")
                # Cleanup flag
                for f_name in ["BATCH_STOP.flag", "PERIODIC_STOP.flag", "update_stop.flag", "ENRICH_STOP.flag"]:
                    try: (scraper_dir / f_name).unlink()
                    except: pass
            else:
                emit_log("ERR", f"Task '{task_name}' failed with exit code {rc}")
                emit_status("error", message=f"Exit code {rc}")
                
        except Exception as e:
            emit_log("ERR", f"Failed to start task: {e}")
        finally:
            update_process = None

    thread = threading.Thread(target=run_and_stream)
    thread.daemon = True
    thread.start()
    
    return jsonify({'status': 'started', 'task': task_name})

@app.route('/api/save-to-bigquery', methods=['POST'])
def save_to_bigquery():
    """Manual trigger to save multiple Excel files' data to BigQuery."""
    data = request.json or {}
    file_paths = data.get('file_paths', [])
    
    if not file_paths:
        return jsonify({'error': 'No files provided'}), 400
        
    import pandas as pd
    total_rows = 0
    successful_files = 0
    
    for file_path in file_paths:
        try:
            emit_log('INFO', f"Uploading {os.path.basename(file_path)} to BigQuery...")
            df = pd.read_excel(file_path)
            # Ensure price/old price are numeric for BigQuery
            for col in ['price', 'old price', 'm2 construidos', 'm2 utiles', 'precio por m2']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            success = db_manager.save_listings_from_df(df, source_file=os.path.basename(file_path))
            
            if success:
                total_rows += len(df)
                successful_files += 1
                emit_log('OK', f"✅ Successfully uploaded {len(df)} rows from {os.path.basename(file_path)}")
            else:
                emit_log('ERR', f"❌ BigQuery upload failed for {os.path.basename(file_path)}")
        except Exception as e:
            emit_log('ERR', f"❌ Error processing {os.path.basename(file_path)}: {e}")
    
    if successful_files > 0:
        return jsonify({
            'status': 'ok', 
            'message': f'Uploaded {total_rows} rows from {successful_files} file(s) to BigQuery'
        })
    else:
        return jsonify({'error': 'No files were successfully uploaded. Check logs.'}), 500

if __name__ == '__main__':
    run_server()
