"""Excel I/O operations for property data export.

This module handles all Excel file operations including:
- Loading existing workbooks to prevent data loss
- Merging new scraping results with existing data
- Preserving non-target worksheets (e.g., keeping 'venta' sheet when updating 'alquiler')
- Applying proper number formatting to numeric fields
- Handling PermissionError when Excel file is open
- Column name normalization for backward compatibility
"""
from __future__ import annotations

import os
from typing import List, Set

import pandas as pd
from openpyxl.utils import get_column_letter

from . import ORDERED_BASE, ORDERED_HABITACIONES
from .utils import log

RENAME_COMPAT = {"plantas":"Num plantas","m2":"Num plantas","description":"Descripción","consumo":"Consumo 1","emisiones":"Emisiones 1"}
FILL_MISSING = ["Consumo 1","Consumo 2","Emisiones 1","Emisiones 2","old price","price change %","gastos comunidad"]


def load_existing_single_sheet(path: str, sheet: str):
    """Load all sheets from an existing Excel file that contain property URLs.
    
    This legacy function unions all sheets containing a 'URL' column to ensure
    no existing data is lost when the scraper updates a specific sheet.
    Column names are normalized for backward compatibility with older exports.
    
    Args:
        path: Path to the Excel file
        sheet: Default sheet name (currently unused, kept for compatibility)
        
    Returns:
        Tuple of (DataFrame, Set[str], Set[str]):
        - DataFrame: Combined data from all sheets with URL columns
        - Set of seen URLs for deduplication
        - Set of column names found in the workbook
        
    Note:
        Returns empty DataFrame with ORDERED_BASE columns if file doesn't exist.
    """
    seen: Set[str] = set()
    cols: Set[str] = set()
    if not os.path.exists(path):
        return pd.DataFrame(columns=ORDERED_BASE), seen, set()

    with pd.ExcelFile(path) as xls:
        frames = []
        for sh in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sh)
            if "URL" not in df.columns:
                continue
            cols_set = set(df.columns)
            df = df.rename(columns={k:v for k,v in RENAME_COMPAT.items() if k in cols_set})
            for c in FILL_MISSING:
                if c not in cols_set:
                    df[c] = None
            frames.append(df)
        if not frames:
            return pd.DataFrame(columns=ORDERED_BASE), seen, set()
        df_all = pd.concat(frames, ignore_index=True)

    cols.update(df_all.columns)
    if "URL" in df_all.columns:
        seen.update(map(str, df_all["URL"].dropna().astype(str)))
    return df_all, seen, cols


def load_urls_with_dates(path: str) -> dict:
    """Load all URLs with their 'actualizado hace' dates from all sheets.
    
    This enables smart deduplication - only re-scrape properties that have been
    updated since the last scrape.
    
    Args:
        path: Path to the Excel file
        
    Returns:
        Dict mapping URL -> last_updated_date (as string)
    """
    url_dates = {}
    if not os.path.exists(path):
        return url_dates
    
    try:
        with pd.ExcelFile(path) as xls:
            for sh in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sh)
                if "URL" not in df.columns:
                    continue
                
                date_col = None
                for col in ["actualizado hace", "actualizado", "last_updated"]:
                    if col in df.columns:
                        date_col = col
                        break
                
                if date_col:
                    for _, row in df.iterrows():
                        url = str(row.get("URL", "")).strip()
                        date = str(row.get(date_col, "")).strip()
                        if url and url != "nan":
                            url_dates[url] = date
                else:
                    # No date column, just mark URLs as existing with empty date
                    for url in df["URL"].dropna().astype(str):
                        url = url.strip()
                        if url and url != "nan":
                            url_dates[url] = ""
    except Exception as e:
        log("WARN", f"Error loading URLs with dates: {e}")
    
    return url_dates

def load_existing_specific_sheet(path: str, sheet: str) -> pd.DataFrame:
    """Load a specific worksheet from an Excel file.
    
    Unlike load_existing_single_sheet, this function only loads the target sheet,
    which is used during final export to merge new data with existing data in
    that specific sheet while preserving other sheets.
    
    Args:
        path: Path to the Excel file
        sheet: Name of the worksheet to load
        
    Returns:
        DataFrame containing the sheet data, or empty DataFrame with ORDERED_BASE
        columns if file/sheet doesn't exist or an error occurs.
        
    Note:
        Column names are normalized for backward compatibility (e.g., 'plantas' → 'Num plantas').
    """
    if not os.path.exists(path):
        return pd.DataFrame(columns=ORDERED_BASE)
    try:
        with pd.ExcelFile(path) as xls:
            if sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
            else:
                return pd.DataFrame(columns=ORDERED_BASE)
    except Exception:
        return pd.DataFrame(columns=ORDERED_BASE)

    cols_set = set(df.columns)
    df = df.rename(columns={k:v for k,v in RENAME_COMPAT.items() if k in cols_set})
    for c in FILL_MISSING:
        if c not in cols_set:
            df[c] = None
    return df

def write_excel_with_retry(df: pd.DataFrame, out_path: str, sheet: str, max_retries: int = 5, check_stop = None):
    """Write a DataFrame to Excel with automatic retry if file is locked.
    
    This function preserves existing worksheets in the workbook - only the target
    sheet is replaced. If the user has the Excel file open, it will retry
    up to max_retries times before giving up.
    
    Args:
        df: DataFrame to write
        out_path: Path to the Excel file
        sheet: Name of the worksheet to write/replace
        max_retries: Maximum number of PermissionError retries
        check_stop: Optional callable that returns True if we should abort
    """
    attempt = 0
    while attempt < max_retries:
        if check_stop and check_stop():
            log("INFO", "Excel write aborted by user/stop signal")
            return
        try:
            mode = "a" if os.path.exists(out_path) else "w"
            with pd.ExcelWriter(
                out_path,
                engine="openpyxl",
                mode=mode,
                if_sheet_exists=("replace" if mode == "a" else None),
            ) as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
                try:
                    ws = writer.book[sheet]
                    header_row = 1
                    n_rows = ws.max_row
                    header_cells = [c.value for c in ws[header_row]]
                    col_index = {name: idx + 1 for idx, name in enumerate(header_cells)}

                    for name in ["Num plantas", "banos", "construido en",
                                 "price", "m2 construidos", "m2 utiles", "precio por m2", "parcela"]:
                        if name in col_index:
                            col = col_index[name]
                            letter = get_column_letter(col)
                            for r in range(header_row + 1, n_rows + 1):
                                cell = ws[f"{letter}{r}"]
                                if isinstance(cell.value, (int, float)):
                                    # Allow 2 decimals for price per m2 if it's small (rent)
                                    if name == "precio por m2" and cell.value < 100:
                                        cell.number_format = "0.00"
                                    else:
                                        cell.number_format = "0"

                    if "price change %" in col_index:
                        col = col_index["price change %"]
                        letter = get_column_letter(col)
                        for r in range(header_row + 1, n_rows + 1):
                            cell = ws[f"{letter}{r}"]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0%"
                except Exception:
                    pass
            break
        except PermissionError:
            attempt += 1
            if attempt >= max_retries:
                log("ERR", f"Failed to write to '{out_path}' after {max_retries} attempts. File is likely open in Excel.")
                raise
            
            log("WARN", f"Cannot write to '{out_path}' (Attempt {attempt}/{max_retries}). The file appears to be open in Excel.")
            log("INFO", "Retrying in 5 seconds... Please close the file.")
            
            import time
            for _ in range(10): # 10 * 0.5s = 5s
                if check_stop and check_stop():
                    return
                time.sleep(0.5)

def export_single_sheet(existing_df: pd.DataFrame,
                        additions: List[dict],
                        out_path: str,
                        sheet: str,
                        carry_cols: Set[str]):
    """Merge new scraping results with existing data and export to Excel.
    
    This is the main export orchestrator that:
    1. Combines existing data with newly scraped properties
    2. Ensures all columns from ORDERED_BASE are present
    3. Adds any extra columns found in the data
    4. Deduplicates by URL (keeps first occurrence)
    5. Converts numeric columns to appropriate data types
    6. Writes to Excel with proper formatting
    
    Args:
        existing_df: Previously scraped data from the target sheet
        additions: List of new property records to add (list of dicts)
        out_path: Path to the output Excel file
        sheet: Name of the worksheet to update
        carry_cols: Additional column names to preserve (currently unused but kept for future)
        
    Side Effects:
        Writes to the Excel file at out_path and logs the result.
    """
    extra_cols = set(carry_cols)
    for r in additions:
        extra_cols.update(r.keys())
    
    # Detect room mode: if m2_habs column is present, use ORDERED_HABITACIONES
    is_room_mode = any('m2_habs' in r for r in additions) or (not existing_df.empty and 'm2_habs' in existing_df.columns)
    base_columns = ORDERED_HABITACIONES if is_room_mode else ORDERED_BASE
    
    ordered = list(base_columns) + [c for c in sorted(extra_cols) if c not in base_columns]

    add_df = pd.DataFrame(additions)

    if existing_df.empty and add_df.empty:
        combined = existing_df.copy()
    elif existing_df.empty:
        combined = add_df.copy()
    elif add_df.empty:
        combined = existing_df.copy()
    else:
        # Drop all-NA columns from both DataFrames before concat to avoid FutureWarning
        existing_clean = existing_df.dropna(axis=1, how='all')
        add_clean = add_df.dropna(axis=1, how='all')
        combined = pd.concat([existing_clean, add_clean], ignore_index=True)

    for c in ordered:
        if c not in combined.columns:
            combined[c] = None

    if "URL" in combined.columns:
        # keep='last' ensures updated rows (from new scrape) replace old ones
        combined = combined.drop_duplicates(subset=["URL"], keep="last")

    int_like_cols = ["Num plantas", "banos", "construido en",
                     "price", "m2 construidos", "m2 utiles", "precio por m2", "parcela"]
    for c in int_like_cols:
        if c in combined.columns:
            combined[c] = pd.to_numeric(combined[c], errors="coerce")
    if "price change %" in combined.columns:
        combined["price change %"] = pd.to_numeric(combined["price change %"], errors="coerce")

    write_excel_with_retry(combined[ordered], out_path, sheet)
    log("OK", f"Saved {len(combined)} total rows -> {out_path} (sheet='{sheet}')")


def export_split_by_distrito(existing_df: pd.DataFrame,
                              additions: List[dict],
                              out_path: str,
                              carry_cols: Set[str],
                              max_retries: int = 5,
                              check_stop = None):
    """Export data split into multiple sheets based on Distrito (column J).
    
    Each unique Distrito value gets its own sheet. This helps organize
    data when scraping multiple districts from a city.
    
    Args:
        existing_df: Previously scraped data
        additions: List of new property records to add
        out_path: Path to the output Excel file
        carry_cols: Additional column names to preserve
        max_retries: Maximum number of PermissionError retries
        check_stop: Optional callable that returns True if we should abort
        
    Side Effects:
        Writes to the Excel file at out_path with multiple sheets.
    """
    extra_cols = set(carry_cols)
    for r in additions:
        extra_cols.update(r.keys())
    ordered = list(ORDERED_BASE) + [c for c in sorted(extra_cols) if c not in ORDERED_BASE]
    
    add_df = pd.DataFrame(additions)
    
    # Combine existing and new data
    if existing_df.empty and add_df.empty:
        combined = existing_df.copy()
    elif existing_df.empty:
        combined = add_df.copy()
    elif add_df.empty:
        combined = existing_df.copy()
    else:
        existing_clean = existing_df.dropna(axis=1, how='all')
        add_clean = add_df.dropna(axis=1, how='all')
        combined = pd.concat([existing_clean, add_clean], ignore_index=True)
    
    for c in ordered:
        if c not in combined.columns:
            combined[c] = None
    
    if "URL" in combined.columns:
        combined = combined.drop_duplicates(subset=["URL"], keep="last")
    
    # Convert numeric columns
    int_like_cols = ["Num plantas", "banos", "construido en",
                     "price", "m2 construidos", "m2 utiles", "precio por m2", "parcela"]
    for c in int_like_cols:
        if c in combined.columns:
            combined[c] = pd.to_numeric(combined[c], errors="coerce")
    if "price change %" in combined.columns:
        combined["price change %"] = pd.to_numeric(combined["price change %"], errors="coerce")
    
    # Debug: log column info
    log("INFO", f"Combined data has {len(combined)} rows, columns include 'Distrito': {'Distrito' in combined.columns}")
    
    # Split by Distrito
    if "Distrito" not in combined.columns or combined.empty:
        # Fallback to single sheet if no Distrito column
        log("WARN", "No Distrito column found, exporting to single sheet")
        write_excel_with_retry(combined[ordered], out_path, "data", max_retries=max_retries, check_stop=check_stop)
        log("OK", f"Saved {len(combined)} rows -> {out_path} (no Distrito column)")
        return
    
    # Get unique Distrito values
    distritos = combined["Distrito"].fillna("Sin Distrito").unique()
    log("INFO", f"Found {len(distritos)} unique Distrito values: {list(distritos)[:5]}...")
    
    attempt = 0
    while attempt < max_retries:
        if check_stop and check_stop():
            log("INFO", "Excel write aborted by user/stop signal")
            return
        try:
            mode = "w"  # Start fresh for multi-sheet export
            with pd.ExcelWriter(
                out_path,
                engine="openpyxl",
                mode=mode,
            ) as writer:
                total_rows = 0
                sheets_created = 0
                for distrito in sorted(distritos):
                    # Filter data for this Distrito
                    if distrito == "Sin Distrito":
                        district_df = combined[combined["Distrito"].isna() | (combined["Distrito"] == "Sin Distrito")]
                    else:
                        district_df = combined[combined["Distrito"] == distrito]
                    
                    if district_df.empty:
                        continue
                    
                    # Sanitize sheet name (Excel has 31 char limit and restrictions)
                    # Also strip "Distrito " prefix if present
                    distrito_str = str(distrito)
                    if distrito_str.lower().startswith("distrito "):
                        distrito_str = distrito_str[9:]  # Remove "Distrito " (9 chars)
                    sheet_name = distrito_str[:31].replace("[", "(").replace("]", ")").replace(":", "-").replace("*", "-").replace("?", "-").replace("/", "-").replace("\\", "-")
                    if not sheet_name:
                        sheet_name = "Sin Distrito"
                    
                    district_df[ordered].to_excel(writer, sheet_name=sheet_name, index=False)
                    total_rows += len(district_df)
                    sheets_created += 1
                    log("INFO", f"Created sheet '{sheet_name}' with {len(district_df)} rows")
                    
                    # Apply number formatting
                    try:
                        ws = writer.book[sheet_name]
                        header_row = 1
                        n_rows = ws.max_row
                        header_cells = [c.value for c in ws[header_row]]
                        col_index = {name: idx + 1 for idx, name in enumerate(header_cells)}
                        
                        for name in ["Num plantas", "banos", "construido en",
                                     "price", "m2 construidos", "m2 utiles", "precio por m2", "parcela"]:
                            if name in col_index:
                                col = col_index[name]
                                letter = get_column_letter(col)
                                for r in range(header_row + 1, n_rows + 1):
                                    cell = ws[f"{letter}{r}"]
                                    if isinstance(cell.value, (int, float)):
                                        # Allow 2 decimals for price per m2 if it's small (rent)
                                        if name == "precio por m2" and cell.value < 100:
                                            cell.number_format = "0.00"
                                        else:
                                            cell.number_format = "0"
                        
                        if "price change %" in col_index:
                            col = col_index["price change %"]
                            letter = get_column_letter(col)
                            for r in range(header_row + 1, n_rows + 1):
                                cell = ws[f"{letter}{r}"]
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = "0%"
                    except Exception as e:
                        log("WARN", f"Error applying formatting to sheet {sheet_name}: {e}")
            
            log("OK", f"Saved {total_rows} rows across {sheets_created} sheets -> {out_path}")
            break
        except PermissionError:
            attempt += 1
            if attempt >= max_retries:
                log("ERR", f"Failed to write to '{out_path}' after {max_retries} attempts.")
                raise
            
            log("WARN", f"Cannot write to '{out_path}' (Attempt {attempt}/{max_retries}). The file appears to be open in Excel.")
            
            import time
            for _ in range(20): # 20 * 0.5s = 10s
                if check_stop and check_stop():
                    return
                time.sleep(0.5)


