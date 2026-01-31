import openpyxl

wb = openpyxl.load_workbook('yield/Calculadora.xlsx', data_only=False)

print("=== GASTOS COMPRA (ITP por Comunidad) ===")
ws = wb['Gastos Compra']
for r in range(1, 25):
    for c in range(1, 5):
        cell = ws.cell(r, c)
        if cell.value is not None:
            val = str(cell.value)[:50]
            print(f"({r:2},{c}): {val}")

print("\n\n=== TRAMOS IRPF ===")
ws2 = wb['Tramos IRPF']
for r in range(1, 15):
    for c in range(1, 5):
        cell = ws2.cell(r, c)
        if cell.value is not None:
            val = str(cell.value)[:50]
            print(f"({r:2},{c}): {val}")
